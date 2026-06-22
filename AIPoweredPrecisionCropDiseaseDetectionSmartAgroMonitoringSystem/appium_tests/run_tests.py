import os
import sys
import json
import time
from datetime import datetime
import pytest

def generate_excel_report(results):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Required libraries (openpyxl) not found. Cannot generate Excel report.")
        print("Please run: pip install pandas openpyxl")
        return False

    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # SHEET 1: TEST RUN SUMMARY
    # ----------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Test Run Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Fonts & Fills
    title_font = Font(name="Segoe UI", size=16, bold=True, color="1B5E20")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    section_font = Font(name="Segoe UI", size=12, bold=True, color="2E7D32")
    normal_font = Font(name="Segoe UI", size=11, color="333333")
    bold_font = Font(name="Segoe UI", size=11, bold=True, color="333333")
    metric_lbl_font = Font(name="Segoe UI", size=10, color="555555")
    metric_val_font = Font(name="Segoe UI", size=12, bold=True, color="1B5E20")
    
    green_header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    green_light_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    gray_light_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    border_side = Side(border_style="thin", color="CCCCCC")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    thick_bottom = Border(bottom=Side(border_style="medium", color="1B5E20"))
    
    # Title Block
    ws_summary.merge_cells("A1:D1")
    ws_summary["A1"] = "AgroAI Precision Farming App E2E Test Report"
    ws_summary["A1"].font = title_font
    ws_summary["A1"].alignment = Alignment(vertical="center")
    ws_summary.row_dimensions[1].height = 40
    
    ws_summary["A2"] = "Automated Appium Quality Assurance Analysis"
    ws_summary["A2"].font = Font(name="Segoe UI", size=11, italic=True, color="666666")
    ws_summary.row_dimensions[2].height = 20
    
    ws_summary["A4"] = "Execution Metadata"
    ws_summary["A4"].font = section_font
    ws_summary["A4"].border = thick_bottom
    
    # Setup Metadata Grid
    metadata = [
        ("Execution Time", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Test Environment", "Local Mobile Emulator"),
        ("Automation Tools", "Appium (UiAutomator2) & pytest"),
        ("Target Package", "com.example.ai_poweredprecisioncropdiseasedetectionsmartagromonitoringsystem"),
        ("MainActivity", ".MainActivity")
    ]
    
    row = 5
    for key, val in metadata:
        ws_summary.cell(row=row, column=1, value=key).font = bold_font
        ws_summary.cell(row=row, column=1).fill = gray_light_fill
        ws_summary.cell(row=row, column=1).border = thin_border
        
        ws_summary.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws_summary.cell(row=row, column=2, value=val).font = normal_font
        ws_summary.cell(row=row, column=2).border = thin_border
        ws_summary.cell(row=row, column=3).border = thin_border
        ws_summary.cell(row=row, column=4).border = thin_border
        row += 1
        
    row += 1
    ws_summary.cell(row=row, column=1, value="Key Performance Indicators (KPIs)").font = section_font
    ws_summary.cell(row=row, column=1).border = thick_bottom
    
    # Calculate Metrics
    total_tests = len(results)
    passed_tests = sum(1 for r in results if r["status"] == "Pass")
    failed_tests = total_tests - passed_tests
    pass_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0.0
    total_duration = sum(r["duration"] for r in results)
    
    row += 1
    kpis = [
        ("Total Runs", total_tests, "0"),
        ("Passed Steps", passed_tests, "0"),
        ("Failed Steps", failed_tests, "0"),
        ("Pass Rate", f"{pass_rate:.1f}%", "@"),
        ("Duration", f"{total_duration:.2f}s", "@")
    ]
    
    kpi_col = 1
    for label, val, num_fmt in kpis:
        # Label cell
        cell_lbl = ws_summary.cell(row=row, column=kpi_col, value=label)
        cell_lbl.font = metric_lbl_font
        cell_lbl.alignment = Alignment(horizontal="center", vertical="center")
        cell_lbl.border = thin_border
        cell_lbl.fill = gray_light_fill
        
        # Value cell
        cell_val = ws_summary.cell(row=row+1, column=kpi_col, value=val)
        cell_val.font = metric_val_font
        cell_val.alignment = Alignment(horizontal="center", vertical="center")
        cell_val.border = thin_border
        cell_val.fill = green_light_fill if label == "Pass Rate" and pass_rate == 100 else gray_light_fill
        if num_fmt != "@":
            cell_val.number_format = num_fmt
            
        kpi_col += 1
        
    ws_summary.row_dimensions[row].height = 25
    ws_summary.row_dimensions[row+1].height = 30
    
    # ----------------------------------------------------
    # SHEET 2: DETAILED TEST LOGS
    # ----------------------------------------------------
    ws_details = wb.create_sheet(title="Detailed Test Logs")
    ws_details.views.sheetView[0].showGridLines = True
    
    # Title
    ws_details.merge_cells("A1:G1")
    ws_details["A1"] = "Detailed End-to-End Test Case Diagnostics"
    ws_details["A1"].font = title_font
    ws_details.row_dimensions[1].height = 35
    
    headers = [
        "Step / Test Case",
        "Description",
        "Status",
        "Duration (s)",
        "Failure Diagnostics / Error Message",
        "Execution Timestamp",
        "Remarks & Quality Analysis"
    ]
    
    # Write Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws_details.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = green_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    ws_details.row_dimensions[3].height = 30
    
    # Status Fills & Fonts
    pass_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    pass_font = Font(name="Segoe UI", size=10, bold=True, color="2E7D32")
    fail_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    fail_font = Font(name="Segoe UI", size=10, bold=True, color="C62828")
    
    # Write Data Rows
    row_idx = 4
    for r in results:
        ws_details.cell(row=row_idx, column=1, value=r["test_name"]).font = bold_font
        ws_details.cell(row=row_idx, column=2, value=r["description"]).font = normal_font
        
        status_cell = ws_details.cell(row=row_idx, column=3, value=r["status"])
        status_cell.alignment = Alignment(horizontal="center")
        if r["status"] == "Pass":
            status_cell.fill = pass_fill
            status_cell.font = pass_font
        else:
            status_cell.fill = fail_fill
            status_cell.font = fail_font
            
        dur_cell = ws_details.cell(row=row_idx, column=4, value=r["duration"])
        dur_cell.font = normal_font
        dur_cell.alignment = Alignment(horizontal="right")
        dur_cell.number_format = "0.00"
        
        err_cell = ws_details.cell(row=row_idx, column=5, value=r["error_msg"])
        err_cell.font = Font(name="Segoe UI", size=9, color="A00000")
        err_cell.alignment = Alignment(wrap_text=True)
        
        ws_details.cell(row=row_idx, column=6, value=r["timestamp"]).font = normal_font
        
        # Add a smart remark based on status
        remark = "Functionality verified, meets performance SLAs." if r["status"] == "Pass" else "Diagnostic review required: view error message details."
        ws_details.cell(row=row_idx, column=7, value=remark).font = normal_font
        
        # Apply borders
        for c in range(1, 8):
            ws_details.cell(row=row_idx, column=c).border = thin_border
            
        ws_details.row_dimensions[row_idx].height = 24
        row_idx += 1
        
    # Autofit Column Widths for both sheets
    for ws in [ws_summary, ws_details]:
        for col in ws.columns:
            # Skip merged cells calculation for column A in row 1
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Avoid title rows skewing widths
                if cell.row in [1, 2] and cell.coordinate in ws.merged_cells:
                    continue
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Save Report
    report_name = "test_analysis_report.xlsx"
    report_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), report_name)
    wb.save(report_path)
    print(f"SUCCESS: Report saved at {report_path}")
    return report_path

def main():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    results_path = os.path.join(current_dir, ".test_results.json")
    
    # Clean up previous runs
    if os.path.exists(results_path):
        os.remove(results_path)
        
    print("====================================================")
    print("STARTING E2E APPIUM TEST RUN FOR AGROAI APPLICATION")
    print("====================================================")
    
    test_file = os.path.join(current_dir, "test_e2e.py")
    
    # Run pytest
    exit_code = pytest.main(["-v", test_file])
    print(f"\nPytest exited with status code: {exit_code}")
    
    # Read collected test results
    results = []
    if os.path.exists(results_path):
        try:
            with open(results_path, "r") as f:
                results = json.load(f)
        except Exception as e:
            print(f"Error reading test results file: {e}")
    else:
        print("WARNING: Temporary results file '.test_results.json' was not found.")
        print("If tests failed to start, check Appium server logs or emulator connection.")
        
    # If no results were logged, create a mock diagnostic entry
    if not results:
        results = [{
            "test_name": "Appium Connection Diagnostics",
            "description": "Checks connectivity to the Appium driver and emulator.",
            "status": "Fail",
            "duration": 0.0,
            "error_msg": "Appium server at http://localhost:4723 was unreachable, or emulator was not running.",
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }]
        
    print("\n====================================================")
    print("GENERATING EXCEL REPORT ANALYSIS")
    print("====================================================")
    report_path = generate_excel_report(results)
    
    print("\nTest Run completed.")
    if report_path:
        print(f"Open your report file here: {report_path}")
    
    sys.exit(exit_code)

if __name__ == "__main__":
    main()
